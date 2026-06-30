"""
prepare_ror_bulk_submission.py
================================
Reads institutions marked 'no_ror_exists' in the curation table, enriches
them with data from the INSPIRE records cache, filters out institutions
likely to be out of ROR scope, and writes the ROR "New Records" bulk
processing spreadsheet ready to email to support@ror.org.

ROR bulk submission process
-----------------------------
  1. Run this script to produce the spreadsheet.
  2. Open the spreadsheet and review/fill in any missing fields — especially
     'Link to publications' (required) which cannot be auto-populated reliably.
  3. Download the official ROR template from:
       https://ror.org/registry/#bulk-requests
     and compare your output columns against it if the template has changed.
  4. Email the completed spreadsheet to support@ror.org with subject:
       "Bulk new record request - Indian HEP institutions"
  5. Track progress on GitHub: https://github.com/ror-community/ror-updates/issues

ROR scope reminder (institutions that will be REJECTED)
---------------------------------------------------------
  - University departments, faculties, schools, colleges within a university
  - Internal units of companies
  - Funding programs/schemes
  - One-person consultancies
  - Organizations not cited as affiliations in published research by multiple people

This script pre-filters on institution_type and paper_count, but you should
review the output manually — especially records with institution_type=[] or
ambiguous names — before submitting.

Usage
------
    # Preview — prints what would be submitted, writes nothing
    python3 prepare_ror_bulk_submission.py

    # Write the spreadsheet
    python3 prepare_ror_bulk_submission.py --output ror_bulk_submission.xlsx

    # Lower the paper-count threshold (default: 5)
    python3 prepare_ror_bulk_submission.py --min-papers 2 --output ror_bulk_submission.xlsx

    # Include institution types that are usually out of scope (for manual review)
    python3 prepare_ror_bulk_submission.py --include-all-types --output ror_bulk_submission.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DB_PATH           = "inspire_ror.db"
RECORDS_CACHE     = "inspire_records_cache.json"
INSPIRE_SEARCH    = "https://inspirehep.net/institutions/{control_number}"

# ROR's valid type values (from the dropdown in the template spreadsheet)
ROR_TYPES = {
    "Archive", "Company", "Education", "Facility",
    "Government", "Healthcare", "Nonprofit", "Other",
}

# Map from INSPIRE institution_type values to ROR type values
INSPIRE_TO_ROR_TYPE: dict[str, str] = {
    "University":                   "Education",
    "Research Center":              "Facility",
    "Research center":              "Facility",
    "Laboratory":                   "Facility",
    "National Laboratory":          "Facility",
    "Government":                   "Government",
    "Company":                      "Company",
    "Archive":                      "Archive",
    "Hospital":                     "Healthcare",
    "Nonprofit":                    "Nonprofit",
    "Other":                        "Other",
}

# Institution types that are explicitly out of ROR scope — excluded by default.
# (These are department/unit-level types, not independent organisations.)
OUT_OF_SCOPE_TYPES = {
    "Department",
    "Group",
    "Section",
    "Division",
}

# Keywords in institution names that suggest out-of-scope sub-units.
# Used as a secondary filter when institution_type is empty or ambiguous.
OUT_OF_SCOPE_NAME_PATTERNS = [
    r"\bdepartment\b",
    r"\bdept\.?\b",
    r"\bdivision\b",
    r"\bsection\b",
    r"\bgroup\b",
    r"\bschool of\b",
    r"\bfaculty of\b",
    r"\bcollege of\b",
]
_OUT_OF_SCOPE_RE = re.compile(
    "|".join(OUT_OF_SCOPE_NAME_PATTERNS), re.IGNORECASE
)


# ---------------------------------------------------------------------------
# DB query
# ---------------------------------------------------------------------------

def get_no_ror_institutions(
    db_path: str,
    min_papers: int,
) -> pd.DataFrame:
    """
    Fetch institutions marked no_ror_exists in the curation table,
    joined with their latest pipeline result and identity data.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    df = pd.read_sql("""
        SELECT
            i.control_number,
            i.legacy_ICN,
            i.official_name,
            i.acronym,
            i.city,
            i.state,
            i.existing_ror,
            i.institution_type,     -- JSON-encoded list
            i.paper_count,
            c.curated_by,
            c.curation_notes
        FROM curation c
        JOIN institutions i ON i.control_number = c.control_number
        WHERE c.curation_status = 'no_ror_exists'
          AND COALESCE(i.paper_count, 0) >= ?
        ORDER BY i.paper_count DESC NULLS LAST
    """, conn, params=(min_papers,))
    conn.close()
    return df


# ---------------------------------------------------------------------------
# INSPIRE cache enrichment
# ---------------------------------------------------------------------------

def load_inspire_cache(cache_path: str) -> dict[int, dict]:
    """Load inspire_records_cache.json into a dict keyed by control_number."""
    p = Path(cache_path)
    if not p.exists():
        print(f"Warning: INSPIRE cache not found at {cache_path}.")
        print("  Run run_pipeline.py first to populate the cache.")
        print("  Fields that come from the cache (URLs, ext_ids, name_variants)")
        print("  will be empty in the output spreadsheet.")
        return {}
    with p.open() as f:
        raw = json.load(f)
    return {r["control_number"]: r for r in raw if r.get("control_number")}


def extract_ext_ids(meta: dict) -> dict[str, str]:
    """Pull GRID, Wikidata, ISNI, Crossref Funder ids from INSPIRE metadata."""
    ids: dict[str, str] = {}
    for ext in meta.get("external_system_identifiers", []):
        schema = (ext.get("schema") or "").upper()
        value  = (ext.get("value") or "").strip()
        if schema == "GRID" and value:
            ids["grid"] = value
        elif schema == "WIKIDATA" and value:
            ids["wikidata"] = f"https://www.wikidata.org/wiki/{value}"
        elif schema == "ISNI" and value:
            ids["isni"] = f"https://isni.org/isni/{value.replace(' ', '')}"
        elif schema == "FUNDREF" and value:
            ids["fundref"] = f"https://doi.org/10.13039/{value}"
    return ids


def extract_url(meta: dict) -> str:
    """Return the first URL from the INSPIRE record."""
    urls = meta.get("urls", [])
    if urls:
        return (urls[0].get("value") or "").strip()
    return ""


def extract_domain(url: str) -> str:
    """Strip scheme and www. to get a bare domain."""
    from urllib.parse import urlparse
    try:
        parsed = urlparse(url if "://" in url else "https://" + url)
        host = (parsed.netloc or parsed.path).lower().split(":")[0]
        return host.removeprefix("www.")
    except Exception:
        return ""


def extract_name_variants(meta: dict) -> str:
    """Return semicolon-joined name_variants from the INSPIRE record."""
    variants = [v.get("value", "") for v in meta.get("name_variants", []) if v.get("value")]
    return "; ".join(variants)


# ---------------------------------------------------------------------------
# Scope filtering
# ---------------------------------------------------------------------------

def parse_institution_types(raw: str | None) -> list[str]:
    """Parse the JSON-encoded institution_type list from the DB."""
    if not raw:
        return []
    try:
        result = json.loads(raw)
        return result if isinstance(result, list) else []
    except (json.JSONDecodeError, TypeError):
        return [raw] if raw else []


def map_to_ror_type(inspire_types: list[str]) -> str:
    """Map INSPIRE institution_type list to the best matching ROR type."""
    for t in inspire_types:
        ror = INSPIRE_TO_ROR_TYPE.get(t)
        if ror:
            return ror
    return ""  # unknown — reviewer must fill in


def is_out_of_scope(
    inspire_types: list[str],
    name: str,
    include_all_types: bool,
) -> tuple[bool, str]:
    """
    Returns (True, reason) if the institution looks out of ROR scope,
    (False, "") otherwise.
    """
    if include_all_types:
        return False, ""

    for t in inspire_types:
        if t in OUT_OF_SCOPE_TYPES:
            return True, f"institution_type='{t}' is typically out of ROR scope"

    if _OUT_OF_SCOPE_RE.search(name or ""):
        return True, f"name suggests a sub-unit (department/faculty/group)"

    return False, ""


# ---------------------------------------------------------------------------
# Build spreadsheet rows
# ---------------------------------------------------------------------------

def build_rows(
    df: pd.DataFrame,
    cache: dict[int, dict],
    include_all_types: bool,
) -> tuple[list[dict], list[dict]]:
    """
    Returns (included_rows, excluded_rows).
    Each row dict has keys matching the ROR template column headers.
    """
    included: list[dict] = []
    excluded: list[dict] = []

    for _, row in df.iterrows():
        cn    = int(row["control_number"])
        name  = (row["official_name"] or row["legacy_ICN"] or "").strip()
        types = parse_institution_types(row.get("institution_type"))

        # Scope filter
        oos, reason = is_out_of_scope(types, name, include_all_types)
        if oos:
            excluded.append({"control_number": cn, "name": name, "reason": reason})
            continue

        # Enrich from cache
        meta       = cache.get(cn, {})
        ext_ids    = extract_ext_ids(meta)
        url        = extract_url(meta)
        domain     = extract_domain(url)
        variants   = extract_name_variants(meta)
        ror_type   = map_to_ror_type(types)
        papers_url = INSPIRE_SEARCH.format(control_number=cn)

        # Acronym: prefer hierarchy acronym, fall back to ICN-extracted
        acronym = (row.get("acronym") or "").strip()

        # City / country
        city    = (row.get("city") or "").strip()
        state   = (row.get("state") or "").strip()
        # ROR wants city, not state — if city is empty but state is known,
        # note it in comments (state may be wrong level of granularity)
        comments = []
        if not city and state:
            comments.append(f"No city in INSPIRE record; state={state}")
        if not ror_type:
            comments.append(f"INSPIRE type(s): {types or 'not specified'} — reviewer must select ROR type")
        if row.get("curation_notes"):
            comments.append(f"Curator note: {row['curation_notes']}")
        comment_str = " | ".join(comments)

        included.append({
            "Organization name*":                               f"{name} (en)",
            "Names in other languages (separate multiples with semicolon)": "",
            "Name variations":                                  variants,
            "Acronym (separate multiples with semicolon)":     f"{acronym} (en)" if acronym else "",
            "Organization website*":                           url,
            "Organization domain*":                            domain,
            "Link to publications associated with this organization*": papers_url,
            "Wikipedia page":                                  "",
            "Wikidata ID":                                     ext_ids.get("wikidata", ""),
            "ISNI ID":                                         ext_ids.get("isni", ""),
            "GRID ID":                                         ext_ids.get("grid", ""),
            "Crossref Funder ID":                              ext_ids.get("fundref", ""),
            "Type of organization*":                           ror_type,
            "Year established":                                "",
            "Parent org in ROR":                               "",
            "Child org in ROR":                                "",
            "Related org in ROR":                              "",
            "City where org is located*":                      city,
            "Country where org is located*":                   "India",
            "Requestor comments":                              comment_str,
            # Extra columns for internal tracking (not in ROR template)
            "_control_number":                                 cn,
            "_paper_count":                                    row.get("paper_count", ""),
            "_INSPIRE_url":                                    papers_url,
        })

    return included, excluded


# ---------------------------------------------------------------------------
# Write spreadsheet
# ---------------------------------------------------------------------------

ROR_TEMPLATE_COLS = [
    "Organization name*",
    "Names in other languages (separate multiples with semicolon)",
    "Name variations",
    "Acronym (separate multiples with semicolon)",
    "Organization website*",
    "Organization domain*",
    "Link to publications associated with this organization*",
    "Wikipedia page",
    "Wikidata ID",
    "ISNI ID",
    "GRID ID",
    "Crossref Funder ID",
    "Type of organization*",
    "Year established",
    "Parent org in ROR",
    "Child org in ROR",
    "Related org in ROR",
    "City where org is located*",
    "Country where org is located*",
    "Requestor comments",
]

INTERNAL_TRACKING_COLS = ["_control_number", "_paper_count", "_INSPIRE_url"]


def write_spreadsheet(
    included: list[dict],
    excluded: list[dict],
    output_path: str,
) -> None:
    """Write the ROR submission spreadsheet plus an excluded-institutions tab."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # ── Sheet 1: Data provided by requestor (matches ROR template) ─────────
    # We load the original template to preserve its data validation (dropdown
    # for Type of organization) rather than creating from scratch, then write
    # our data rows starting at row 3 (rows 1–2 are Field/Guidance in template).
    template_path = "/mnt/user-data/uploads/PUBLIC_ROR_Bulk_Processing_Template_-_New_Records.xlsx"
    if Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = wb["Data provided by requestor"]
        # Clear example row (row 3 in template = index 2 in 1-based openpyxl)
        for col in range(1, len(ROR_TEMPLATE_COLS) + len(INTERNAL_TRACKING_COLS) + 1):
            ws.cell(row=3, column=col).value = None
    else:
        # Fallback: build from scratch if template not available in this env
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data provided by requestor"
        # Write header
        for col_idx, col_name in enumerate(ROR_TEMPLATE_COLS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

    # Add internal tracking columns header in the first row (after ROR cols)
    tracking_start_col = len(ROR_TEMPLATE_COLS) + 1
    for i, col_name in enumerate(INTERNAL_TRACKING_COLS):
        cell = ws.cell(row=1, column=tracking_start_col + i, value=col_name)
        cell.font = Font(bold=True, italic=True, color="808080")

    # Write data rows
    DATA_START_ROW = 3
    for row_idx, row in enumerate(included, start=DATA_START_ROW):
        for col_idx, col_name in enumerate(ROR_TEMPLATE_COLS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
        for i, col_name in enumerate(INTERNAL_TRACKING_COLS):
            ws.cell(row=row_idx, column=tracking_start_col + i, value=row.get(col_name, ""))

    # Column widths
    col_widths = {
        1:  45,  # Organization name
        5:  35,  # Website
        6:  25,  # Domain
        7:  50,  # Publications link
        13: 15,  # Type
        19: 15,  # City
        20: 10,  # Country
        21: 40,  # Comments
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: Excluded institutions (internal, not sent to ROR) ─────────
    if "Excluded (not submitted)" in wb.sheetnames:
        del wb["Excluded (not submitted)"]
    ws2 = wb.create_sheet("Excluded (not submitted)")

    excl_headers = ["control_number", "name", "reason"]
    for col_idx, h in enumerate(excl_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
    for row_idx, row in enumerate(excluded, start=2):
        for col_idx, h in enumerate(excl_headers, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 60

    # ── Sheet 3: Instructions ───────────────────────────────────────────────
    if "Submission instructions" in wb.sheetnames:
        del wb["Submission instructions"]
    ws3 = wb.create_sheet("Submission instructions")
    instructions = [
        ("BEFORE SUBMITTING — review and complete these required fields:", True),
        ("", False),
        ("1. Organization name*", True),
        ("   Check that the name is the full official English name, not an abbreviation.", False),
        ("   Format: 'Name (en)'", False),
        ("", False),
        ("2. Organization website*", True),
        ("   Must be a working URL. If blank, find the current website manually.", False),
        ("", False),
        ("3. Organization domain*", True),
        ("   The bare domain used for email/web (e.g. iitb.ac.in). No http:// prefix.", False),
        ("", False),
        ("4. Link to publications*  ← MOST IMPORTANT", True),
        ("   The INSPIRE URL pre-filled here is good evidence, but ROR wants a page showing", False),
        ("   the organization's name appearing as an affiliation in published research.", False),
        ("   Ideal: the INSPIRE institution page URL already pre-filled shows this.", False),
        ("", False),
        ("5. Type of organization*", True),
        ("   Must be one of: Archive, Company, Education, Facility, Government,", False),
        ("   Healthcare, Nonprofit, Other.", False),
        ("   Rows with empty Type need you to fill in the correct value.", False),
        ("", False),
        ("6. City where org is located*", True),
        ("   Must be a city name, not a state name. Check rows where city is blank.", False),
        ("", False),
        ("SUBMISSION", True),
        ("  Email completed spreadsheet to: support@ror.org", False),
        ("  Subject: 'Bulk new record request - Indian HEP institutions'", False),
        ("  Track progress: https://github.com/ror-community/ror-updates/issues", False),
        ("", False),
        ("SCOPE REMINDER — ROR will reject:", True),
        ("  - University departments, faculties, schools, colleges", False),
        ("  - Internal company units", False),
        ("  - Organizations NOT cited as affiliations in published research", False),
        ("  See 'Excluded (not submitted)' sheet for institutions filtered out.", False),
        ("", False),
        ("NOTE: One form per organization if submitting via web form.", True),
        ("      This spreadsheet is for the bulk email route (support@ror.org).", False),
    ]
    for row_idx, (text, bold) in enumerate(instructions, start=1):
        cell = ws3.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold)
    ws3.column_dimensions["A"].width = 90

    wb.save(output_path)
    print(f"Spreadsheet written: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Prepare ROR bulk new-record submission from no_ror_exists institutions",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--output", "-o", default=None,
                        help="Output .xlsx path. If omitted, preview only (no file written).")
    parser.add_argument("--min-papers", type=int, default=5,
                        help="Only include institutions with >= this many papers (default: 5).")
    parser.add_argument("--db", default=DB_PATH,
                        help=f"Path to inspire_ror.db (default: {DB_PATH})")
    parser.add_argument("--cache", default=RECORDS_CACHE,
                        help=f"Path to inspire_records_cache.json (default: {RECORDS_CACHE})")
    parser.add_argument("--include-all-types", action="store_true",
                        help="Include institutions that look like sub-units (departments etc) "
                             "instead of filtering them out. Use for manual review.")
    args = parser.parse_args()

    # Load data
    print(f"Reading DB: {args.db}")
    df = get_no_ror_institutions(args.db, args.min_papers)
    print(f"  {len(df)} institutions with curation_status='no_ror_exists' and "
          f"paper_count >= {args.min_papers}")

    if df.empty:
        print("Nothing to submit. Either no institutions are marked 'no_ror_exists', "
              "or all are below the --min-papers threshold.")
        return

    print(f"\nLoading INSPIRE cache: {args.cache}")
    cache = load_inspire_cache(args.cache)
    print(f"  {len(cache)} records in cache")

    # Build rows
    included, excluded = build_rows(df, cache, args.include_all_types)

    # Summary
    print(f"\n{'=' * 60}")
    print(f"  Included in submission : {len(included)}")
    print(f"  Excluded (out of scope): {len(excluded)}")
    print(f"{'=' * 60}")

    # Preview
    if included:
        print("\nInstitutions that WILL be submitted:")
        print(f"  {'Papers':>6}  {'Control#':>10}  Name")
        print(f"  {'-'*6}  {'-'*10}  {'-'*45}")
        for row in included:
            print(f"  {row['_paper_count']:>6}  {row['_control_number']:>10}  "
                  f"{row['Organization name*'][:50]}")

    if excluded:
        print("\nInstitutions EXCLUDED (out of scope):")
        for row in excluded:
            print(f"  {row['control_number']:>10}  {row['name'][:40]:<40}  → {row['reason']}")

    # Check for rows needing manual attention
    needs_attention = [
        r for r in included
        if not r["Organization website*"]
        or not r["Type of organization*"]
        or not r["City where org is located*"]
    ]
    if needs_attention:
        print(f"\n⚠  {len(needs_attention)} row(s) need manual completion before submitting:")
        for r in needs_attention:
            missing = []
            if not r["Organization website*"]:
                missing.append("website")
            if not r["Type of organization*"]:
                missing.append("type")
            if not r["City where org is located*"]:
                missing.append("city")
            print(f"   {r['_control_number']:>10}  {r['Organization name*'][:40]:<40}  "
                  f"missing: {', '.join(missing)}")

    # Write
    if args.output:
        write_spreadsheet(included, excluded, args.output)
        print(f"\nNext steps:")
        print(f"  1. Open {args.output} and review/complete all rows.")
        print(f"  2. Fill in any missing websites, types, or cities.")
        print(f"  3. Verify 'Link to publications' shows the org as an affiliation.")
        print(f"  4. Email to support@ror.org with subject:")
        print(f"     'Bulk new record request - Indian HEP institutions'")
        print(f"  5. Track on GitHub: https://github.com/ror-community/ror-updates/issues")
    else:
        print("\n(Preview only — pass --output <filename.xlsx> to write the spreadsheet)")


if __name__ == "__main__":
    main()