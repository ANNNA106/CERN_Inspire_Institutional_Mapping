from openpyxl import load_workbook
import pandas as pd

COLOR_MAP = {
    0: "accepted",
    9: "review",
    2: "wrong",
    5: "unsure"
}

wb = load_workbook("review_queue_tier3_no_candidate.xlsx")
ws = wb.active

data = []

headers = [c.value for c in ws[1]]

for row in ws.iter_rows(min_row=2):
    values = [c.value for c in row]

    theme = row[0].fill.fgColor.theme
    category = COLOR_MAP.get(theme, "unknown")

    values.append(category)
    data.append(values)

headers.append("review_category")

df = pd.DataFrame(data, columns=headers)
df.to_csv("review_queue_with_categories.csv", index=False)